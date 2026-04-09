import logging
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from margin.models import MarginItem, ProcessedRecord

logger = logging.getLogger(__name__)


def _init_db(db_path: str) -> sqlite3.Connection:
    """Initialize the tracker database and return a connection."""
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS processed (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account_number TEXT NOT NULL,
            error_type TEXT NOT NULL,
            margin_date TEXT NOT NULL,
            dollar_amount TEXT NOT NULL,
            jira_ticket_key TEXT,
            email_sent INTEGER DEFAULT 0,
            processed_at TEXT NOT NULL,
            run_id TEXT NOT NULL,
            flagged_duplicate INTEGER DEFAULT 0,
            advisor_firm TEXT,
            advisor_email TEXT
        )
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_account_error
        ON processed(account_number, error_type)
    """)
    conn.commit()

    # Add new columns to existing DBs (no-op if they exist)
    try:
        conn.execute("ALTER TABLE processed ADD COLUMN advisor_firm TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass  # Column already exists

    try:
        conn.execute("ALTER TABLE processed ADD COLUMN advisor_email TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass  # Column already exists

    return conn


class Tracker:
    def __init__(self, db_path: str):
        self.conn = _init_db(db_path)

    def check_duplicate(self, item: MarginItem) -> Optional[ProcessedRecord]:
        """Check if this account+error combo was already processed."""
        cursor = self.conn.execute(
            """
            SELECT id, account_number, error_type, margin_date, dollar_amount,
                   jira_ticket_key, email_sent, processed_at, run_id, flagged_duplicate
            FROM processed
            WHERE account_number = ? AND error_type = ? AND flagged_duplicate = 0
            ORDER BY processed_at DESC LIMIT 1
            """,
            (item.account_number, item.error_type),
        )
        row = cursor.fetchone()
        if row is None:
            return None
        return ProcessedRecord(
            id=row[0],
            account_number=row[1],
            error_type=row[2],
            margin_date=row[3],
            dollar_amount=row[4],
            jira_ticket_key=row[5],
            email_sent=bool(row[6]),
            processed_at=row[7],
            run_id=row[8],
            flagged_duplicate=bool(row[9]),
        )

    def record_processing(
        self,
        item: MarginItem,
        run_id: str,
        jira_ticket_key: Optional[str] = None,
        email_sent: bool = False,
        advisor_firm: Optional[str] = None,
        advisor_email: Optional[str] = None,
    ) -> None:
        """Record a successfully processed item."""
        self.conn.execute(
            """
            INSERT INTO processed
                (account_number, error_type, margin_date, dollar_amount,
                 jira_ticket_key, email_sent, processed_at, run_id, flagged_duplicate,
                 advisor_firm, advisor_email)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
            """,
            (
                item.account_number,
                item.error_type,
                str(item.date),
                str(item.dollar_amount),
                jira_ticket_key,
                int(email_sent),
                datetime.now().isoformat(),
                run_id,
                advisor_firm,
                advisor_email,
            ),
        )
        self.conn.commit()

    def flag_duplicate(self, item: MarginItem, run_id: str) -> None:
        """Record a flagged duplicate for human review."""
        self.conn.execute(
            """
            INSERT INTO processed
                (account_number, error_type, margin_date, dollar_amount,
                 jira_ticket_key, email_sent, processed_at, run_id, flagged_duplicate)
            VALUES (?, ?, ?, ?, NULL, 0, ?, ?, 1)
            """,
            (
                item.account_number,
                item.error_type,
                str(item.date),
                str(item.dollar_amount),
                datetime.now().isoformat(),
                run_id,
            ),
        )
        self.conn.commit()

    def update_jira_key(self, item: MarginItem, run_id: str, jira_key: str) -> None:
        """Update the Jira ticket key for an existing record."""
        self.conn.execute(
            """
            UPDATE processed SET jira_ticket_key = ?
            WHERE account_number = ? AND error_type = ? AND run_id = ? AND flagged_duplicate = 0
            """,
            (jira_key, item.account_number, item.error_type, run_id),
        )
        self.conn.commit()

    def update_email_sent(self, item: MarginItem, run_id: str) -> None:
        """Mark that the email was sent for this item."""
        self.conn.execute(
            """
            UPDATE processed SET email_sent = 1
            WHERE account_number = ? AND error_type = ? AND run_id = ? AND flagged_duplicate = 0
            """,
            (item.account_number, item.error_type, run_id),
        )
        self.conn.commit()

    def get_jira_key(self, item: MarginItem) -> Optional[str]:
        """Get the Jira ticket key for a previously processed item."""
        cursor = self.conn.execute(
            """
            SELECT jira_ticket_key FROM processed
            WHERE account_number = ? AND error_type = ? AND jira_ticket_key IS NOT NULL
            ORDER BY processed_at DESC LIMIT 1
            """,
            (item.account_number, item.error_type),
        )
        row = cursor.fetchone()
        return row[0] if row else None

    def get_flagged_duplicates(self) -> List[ProcessedRecord]:
        """Return all flagged duplicates."""
        cursor = self.conn.execute(
            """
            SELECT id, account_number, error_type, margin_date, dollar_amount,
                   jira_ticket_key, email_sent, processed_at, run_id, flagged_duplicate
            FROM processed
            WHERE flagged_duplicate = 1
            ORDER BY processed_at DESC
            """
        )
        return [
            ProcessedRecord(
                id=row[0],
                account_number=row[1],
                error_type=row[2],
                margin_date=row[3],
                dollar_amount=row[4],
                jira_ticket_key=row[5],
                email_sent=bool(row[6]),
                processed_at=row[7],
                run_id=row[8],
                flagged_duplicate=bool(row[9]),
            )
            for row in cursor.fetchall()
        ]

    def clear_flagged(self) -> int:
        """Clear all flagged duplicates after review. Returns count deleted."""
        cursor = self.conn.execute("DELETE FROM processed WHERE flagged_duplicate = 1")
        self.conn.commit()
        return cursor.rowcount

    def write_history_sheet(self, file_path: str) -> None:
        """Write processed history to Sheet 2 of the margin Excel file."""
        if not Path(file_path).exists():
            logger.warning(f"Margin file not found, skipping history write: {file_path}")
            return

        try:
            wb = load_workbook(file_path)
        except Exception as e:
            logger.error(f"Failed to load workbook {file_path}: {e}")
            return

        # Remove existing History sheet
        if "History" in wb.sheetnames:
            del wb["History"]

        # Create new History sheet
        ws = wb.create_sheet("History", 1)

        # Headers
        headers = [
            "Account #",
            "Firm",
            "Advisor Email",
            "Error Type",
            "Date",
            "Amount",
            "Jira Ticket",
            "Email Sent",
            "Processed At",
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Query all processed records (non-flagged)
        cursor = self.conn.execute(
            """
            SELECT account_number, advisor_firm, advisor_email, error_type,
                   margin_date, dollar_amount, jira_ticket_key, email_sent, processed_at
            FROM processed
            WHERE flagged_duplicate = 0
            ORDER BY processed_at DESC
            """
        )

        # Write rows
        for row_num, row in enumerate(cursor.fetchall(), 2):
            ws.cell(row=row_num, column=1, value=row[0])  # account_number
            ws.cell(row=row_num, column=2, value=row[1])  # advisor_firm
            ws.cell(row=row_num, column=3, value=row[2])  # advisor_email
            ws.cell(row=row_num, column=4, value=row[3])  # error_type
            ws.cell(row=row_num, column=5, value=row[4])  # margin_date
            ws.cell(row=row_num, column=6, value=row[5])  # dollar_amount
            ws.cell(row=row_num, column=7, value=row[6])  # jira_ticket_key
            ws.cell(row=row_num, column=8, value="Yes" if row[7] else "No")  # email_sent
            ws.cell(row=row_num, column=9, value=row[8])  # processed_at

        try:
            wb.save(file_path)
            logger.info(f"Wrote history sheet to {file_path}")
        except Exception as e:
            logger.error(f"Failed to save history sheet to {file_path}: {e}")

    def close(self):
        self.conn.close()
